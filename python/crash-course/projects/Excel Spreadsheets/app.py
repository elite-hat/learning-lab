import openpyxl as xl

wb = xl.load_workbook("transactions.xlsx")
sheet = wb["Sheet1"]
cell = sheet.cell(1, 1)         # It is equal to: `sheet["a1"]`
print(cell.value)
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 1)